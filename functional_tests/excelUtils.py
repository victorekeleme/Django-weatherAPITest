import openpyxl



def getRowCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.max_row


def getColumnCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.max_column


def readData(file, sheetname, rowNumber, colNumber):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rowNumber, column=colNumber).value


def writeData(file, sheetname, rowNumber, colNumber, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rowNumber, column=colNumber).value = data
    workbook.save(file)
